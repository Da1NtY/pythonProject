import openpyxl
import pymysql

# 从数据库某张表中取出所有记录的函数
# 参数host指定数据库服务器的IP地址
# 参数db_name指定数据库名
# 参数table_name指定表名
# 参数user指定数据库的登录用户名
# 参数password指定登录用户的密码
def get_rows(host,db_name,table_name,user,password):
    conn = pymysql.connect(host=host, port=3306, database=db_name, user=user, password=password)

    # 建立一个游标
    cur = conn.cursor()

    # 组合一个SQL查询语句
    sql = 'select * from '+table_name

    # 执行SQL语句
    cur.execute(sql)

    # rows获得所有行记录，cur.filechall()返回所有符合条件的记录
    rows = cur.fetchall()
    # cur.description返回数据表的字段信息
    # 返回值fields是一个元组，其中的每一项元素也是一个元组（子元组）
    # 这个子元组的第一个元素是字段名
    fields = cur.description
    # 关闭游标
    cur.close()

    # 断开连接
    conn.close()
    return fields,rows

# 将表的记录导入Excel表的函数中
# 参数host指定数据库服务器的IP地址
# 参数db_name指定数据库名
# 参数table_name指定表名
# 参数user指定数据库的登录用户名
# 参数password指定登录用户的密码
# 参数filename指定导入的Excel文件名
def export_to_xlsx(host,db_name,table_name,user,password,filename):
    # 调用函数，取得数据表的字段信息和记录信息
    fields,table_rows = get_rows(host,db_name,table_name,user,password)

    # 生成Excel文件的工作簿
    workbook = openpyxl.Workbook()
    # 在工作簿中生产一个工资表，表名设为“table_”+数据表名
    sheet = workbook.create_sheet('table_' + table_name,0)
    # 在工作表的第一行写上字段名
    for i in range(len(fields)):
        # 在openpyxl模块中定义工作表的行起始值是1，列起始值是1，所以cell()函数第一个参数是1表示第1行，
        # 第二个参数为i+1是因为i从0开始计数，fields[i][0]获得字段的名称
        sheet.cell(1,i+1,fields[i][0])
        # 从工作表第二行开始写入每条记录的内容
        for row in range(len(table_rows)):
            for col in range(len(fields)):
                sheet.cell(row+2, col+1, '%s' % table_rows[row][col])

        # 保存到Excel文件中
        workbook.save(filename)

# 主函数
if __name__ == '__main__':
    # 初始化各变量名
    host = 'localhost'
    db_name = 'test'
    table_name = 'myapp_userinfo'
    user = 'root'
    password = '123456'

    # 调用函数，将一张数据表的全部记录导入到一个Excel文件中
    export_to_xlsx(host,db_name,table_name,user,password,'./userinfo.xlsx')